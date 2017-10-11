# -*- coding: utf-8 -*-
"""
Created on Wed Sep 27 10:04:59 2017

@author: Maddox.Meng
"""

'''L-line'''

import os
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
#from openpyxl.utils.dataframe import dataframe_to_columns

#from  datetime  import  * 
import datetime 
import  time  

def stringToDate(string): 
    #example '2013-07-22 09:44:15+00:00' 
    dt = datetime.datetime.strptime(string, "%Y-%m-%d") 
    #print dt 
    return dt

def intToDays(year):
    return datetime.timedelta(days=365*int(year))


def month(date):
    return int(date.split('-')[1])

def year(date):
    return int(date.split('-')[0])

#if datetime.date(2016,12,31) > stringToDate('2015-12-31'):
#    print 'yes'

path = 'C:\Workspace\AuditAutomation_L\L_workflow'
#CYstart = datetime.date(2016,12,31)#'2016-12-31'
#CYend   = datetime.date(2015,12,31)#'2015-12-31'
CYstart = stringToDate('2015-12-31')#'2016-12-31'
CYend   = stringToDate('2016-12-31')#'2015-12-31'
print CYstart + datetime.timedelta(days=365*5)

for root, dirs, files in os.walk(path):
    for file in files:
        if file.decode('gbk').find('~$') == -1 and file.decode('gbk').find('PBC') != -1: # eliminate temp excel files
            PBC = os.path.join(root,file.decode('gbk'))
        if file.decode('gbk').find('~$') == -1 and file.decode('gbk').find('WP') != -1:
            WP = os.path.join(root,file.decode('gbk'))
            
#print PBC.encode('utf-8')
#print WP.encode('utf-8') 

#wb = load_workbook(WP)
#for sheetname in wb.sheetnames:
#    print sheetname
#    ws = wb.get_sheet_by_name(sheetname)
    
    
    
xl = pd.ExcelFile(PBC)

wb = load_workbook(WP)

for sheet_name in xl.sheet_names:#PBC的每个sheet L110-无形资产清单  L120-无形资产处置清单 
    print sheet_name
    df = pd.read_excel(open(PBC,'rb'), sheetname=sheet_name, header = None)
    
    #找到“序号” 
    for x in range(0, df.shape[0]):
        for y in range(0, df.shape[1]):
            if df.iloc[x,y] == u'序号':
                x_min = x
                y_min = y
                
                
    #df_rel = df.iloc[x_min:df.shape[0], y_min:df.shape[1]] 
    #设置“序号”那行为header           
    df_rel = pd.read_excel(open(PBC,'rb'), sheetname=sheet_name, header = x_min)
    
    '''添加计算列'''
    #df_rel['sum'] = df_rel.apply(lambda x: x.sum(), axis=1)
    #df_rel['净值'] = df_rel.apply(lambda row: row[u'原始成本']-row[u'累计摊销额'], axis=1)
    
    df_rel[u'净值']           = df_rel.apply(lambda row: row[u'原始成本']-row[u'累计摊销额'], axis=1)
    df_rel[u'本年新增标志']    = df_rel.apply(lambda row: 1 if stringToDate(row[u'开始摊销时间']) > CYstart else 0, axis=1) # and (stringToDate(row[u'开始摊销时间']) < CYend
    df_rel[u'年初摊销结束标志']= df_rel.apply(lambda row: 1 if (stringToDate(row[u'开始摊销时间'])  +  intToDays(row[u'总摊销年限'])) > CYstart else 0, axis=1)
    df_rel[u'年末摊销结束标志']= df_rel.apply(lambda row: 1 if (stringToDate(row[u'开始摊销时间'])  +  intToDays(row[u'总摊销年限'])) > CYend else 0, axis=1)
    df_rel[u'每月摊销金额']    = df_rel.apply(lambda row: row[u'原始成本']/row[u'总摊销年限']*12, axis=1)
    
    df_rel[u'本年摊销月份']    = df_rel.apply(lambda row: 13-month(row[u'开始摊销时间']) if row[u'本年新增标志'] == 1 else (0 if (row[u'本年新增标志'] == 0 and row[u'年初摊销结束标志'] == 1) else (13-month(row[u'开始摊销时间']+row[u'总摊销年限'] if (row[u'本年新增标志'] == 0 and row[u'年末摊销结束标志'] == 1) else 12)) ,axis=1)
    
#    df_rel[u'累计摊销月份']    = df_rel.apply(lambda row: min(CYend - stringToDate(row[u'开始摊销时间'])/30, ,row[u'总摊销年限']*12), axis=1)
#    df_rel[u'EY累计摊销额']    = df_rel.apply(lambda row: row[u'每月摊销金额']*row[u'累计摊销月份'], axis=1)
#    df_rel[u'DIFF']           = df_rel.apply(lambda row: row[u'EY累计摊销额']*row[u'累计摊销额'], axis=1)
#    df_rel[u'本年摊销金额']    = df_rel.apply(lambda row: row[u'每月摊销金额']*row[u'本年摊销月份'], axis=1)
#    df_rel[u'新增抽样标志']    = df_rel.apply(lambda row: row[u'每月摊销金额']*row[u'本年摊销月份'], axis=1)
    
    
    
    '''写入WP'''
    '''我决定最佳方案是按列写入，所以先按列读取数据'''
    #遍历headers，get每个header对应的数据，即某一列数据
    for header in list(df_rel):
        #print header
        #print df_rel[header]
        pass

         
     #在WP众多sheets里选出带LL110、L120的sheet
    for sheetname_WP in wb.sheetnames:
        
        if sheetname_WP in sheet_name:
            print sheetname_WP, "in", sheet_name
            ws = wb.get_sheet_by_name(sheetname_WP)  #打开它为ws
        
    for x in range(1, ws.max_row+1):
        for y in range(0, ws.max_column):
            if ws[x][y].value == u'序号':
                print x,y,ws[x][y].value
                rownumber = x
                
    for header in list(df_rel):            
        for cell in ws[rownumber]:
            if cell.value == header:
                
                print 'cell.vaue = header =',cell.value,cell.row,cell.column
                data_col = df_rel[header]
                data_wp_col = column_index_from_string(cell.column) - 1
                
                i=0
                y = data_wp_col
                #for x in range(rownumber+1, ws.max_row+1):
                for x in range(rownumber+1, rownumber+1+len(data_col)):    
                    
                    ws[x][y].value = data_col[i]
                    
                    i+=1
                    
                    
#                i=rownumber+1    
#                for cell in ws[data_wp_col]:
#                    if i >= rownumber + 1 and (i - rownumber -1) < len(data_col):
#                        print 'i=',i
#                        cell.value = data_col[i-rownumber-1]  #1st x is 8
#                    i+=1
#wb.save(WP)           

#wb0 = load_workbook(PBC)
#ws = wb0.get_sheet_by_name(u'L120-无形资产处置清单')
#for cell in ws[11]:
#    print cell.value, cell.coordinate


wbt = load_workbook('tst.xlsx')
wst = wbt.get_active_sheet()
wst[1][0].value = 5
wst[1][1].value = 7
wst[1][2].value = '=A1+B1'
wbt.save('tst.xlsx')                


list1 = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
list2 = [u'序号', u'资产编号', u'无形资产项目', u'无形资产分类', u'总摊销年限', u'开始摊销时间', u'原始成本', u'累计摊销额']
list3 = [u'净值', u'本年新增标志', u'年初摊销结束标志', u'年末摊销结束标志', u'每月摊销金额', u'本年摊销月份', u'累计摊销月份', u'EY累计摊销额', u'DIFF', u'本年摊销金额', u'新增抽样标志']

for i in list3:
    for cell in ws[10]:
        if cell.value == i:
            print 'xxxx',cell.value, cell.column

for cell in ws[10]:
    print cell.value
    


#B'序号', C'资产编号', D'无形资产项目', E'无形资产分类', 
#F'总摊销年限', G'开始摊销时间', H'原始成本', I'累计摊销额'
dict = {}
k=0
for i in list1:
    dict[k] = {'index': i, 'title': list2[k]}
    k+=1

s = u'=原始成本%s-累计摊销额%s'.encode('utf-8') % ('8', '8')
                
'''
=原始成本%s-累计摊销额%s
=IF(开始摊销时间8*1>BG!$B$5,1,0)
=IF((开始摊销时间8+总摊销年限8*365)<($B$4-365),1,0)
=IF(年初摊销结束标志8=1,0,IF((开始摊销时间8+总摊销年限8*365)<$B$4,1,0))
=原始成本8/(总摊销年限8*12)
=IF(年初摊销结束标志8=1,0,IF(本年新增标志8=1,13-MONTH(开始摊销时间8),IF(年末摊销结束标志8=1,MONTH(G开始摊销时间),12)))
=MIN((YEAR($B$4)-YEAR(开始摊销时间8))*12+13-MONTH(开始摊销时间8),总摊销年限8*12)
=累计摊销月份8*每月摊销金额8
=EY累计摊销额8-累计摊销额8
=每月摊销金额8*本年摊销月份8
=IF(AND(本年新增标志8=1,原始成本8>(BG!B7*'L-100'!C64)),"Y","")



净值=            原始成本-累计摊销额
本年新增标志=     if 开始摊销时间 in 本年年初~本年年末 then 1
	              else 0
年初摊销结束标志= if 开始摊销时间+总摊销年限>本年年初 then 1
	              else 0
年末摊销结束标志= if 开始摊销时间+总摊销年限>本年年末 then 1
	              else 0
每月摊销金额=     原始成本/总摊销年限*12
本年摊销月份=     if (本年新增标志 =1) then 13-month(开始摊销时间)
	              else if (本年新增标志 =0 &&年初摊销结束标志=1) then 0
	              else if (本年新增标志 =0 &&年末摊销结束标志=1) then 13	                 -month(开始摊销时间+总摊销年限)
	              else 12
累计摊销月份=    min(month(本年年末-开始摊销时间+1),总摊销年限*12)
EY累计摊销额=    每月摊销金额*累计摊销月份
DIFF=           EY累计摊销额-累计摊销额
本年摊销金额=    每月摊销金额*本年摊销月份
新增抽样标志=    if 本年新增标志=1&&阈值&抽样比例(pending)
'''                
                
            