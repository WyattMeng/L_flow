# -*- coding: utf-8 -*-
"""
Created on Wed Oct 11 14:22:36 2017

@author: Maddox.Meng
"""

import os
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string

#from  datetime  import  * 
import datetime 
#import  time  

def stringToDate(string): # datetime.datetime(2016, 12, 31, 0, 0)
    #example '2013-07-22 09:44:15+00:00' 
    if string.find('-') != -1:
        dt = datetime.date.strptime(string, "%Y-%m-%d") 
    elif string.find('/') != -1:
        dt = datetime.date.strptime(string, "%Y/%m/%d")
    else:
        dt = datetime.date.strptime(string, "%Y-%m-%d")
    return dt

def intToDays(year):
    return datetime.timedelta(days=365*int(year))

def month(date):
    if date.find('-') != -1:
        return int(date.split('-')[1])
    elif date.find('/') != -1:
        return int(date.split('/')[1])

def year(date):
    if date.find('-') != -1:
        return int(date.split('-')[0])
    elif date.find('/') != -1:
        return int(date.split('/')[0])

def monthTotal(date):
    return year(date)*12 + month(date)

'''处理当pandas读到日期格式为pandas._libs.tslib.Timestamp的cell，转成string（）'''
def convtPdTimeToStr(time): # datetime.date(2012, 9, 18)
    if isinstance(time, pd._libs.tslib.Timestamp) is True:
        dateStr = time.date().strftime('%Y-%m-%d')
    elif isinstance(time, unicode) is True or isinstance(time, str):
        dateStr = time
    else:
        dateStr = 'date type is %s' % type(time)
    return dateStr  

def convtPdTimeToDate(time):
    if isinstance(time, pd._libs.tslib.Timestamp) is True:
        dateStr = time.date()
    elif isinstance(time, unicode) is True or isinstance(time, str):
        dateStr = stringToDate(time)
    else:
        dateStr = 'date type is %s' % type(time)
    return dateStr   

cyStart = datetime.date(2015, 12, 31)#stringToDate('2015-12-31')#'2016-12-31'
cyStartStr = '2015-12-31'
cyEnd   = datetime.date(2016, 12, 31)#stringToDate('2016-12-31')#'2015-12-31'
cyEndStr = '2016-12-31'


PBC = 'PBCtst.xlsx'

xl = pd.ExcelFile(PBC)
df = pd.read_excel(open(PBC,'rb'), sheetname='Sheet1', header = None)
#df = pd.read_excel(PBC, header = None)

#找到“序号” 
for x in range(0, df.shape[0]):
    for y in range(0, df.shape[1]):
        if df.iloc[x,y] == u'序号':
            x_min = x
            y_min = y
            
            
#设置“序号”那行为header           
#df_rel = pd.read_excel(PBC, header = x_min)
df_rel = pd.read_excel(open(PBC,'rb'), sheetname='Sheet1', header = x_min)            

df_rel[u'净值']           = df_rel.apply(lambda row: row[u'原始成本']-row[u'累计摊销额'], axis=1)
df_rel[u'本年新增标志']    = df_rel.apply(lambda row: 1 if convtPdTimeToDate(row[u'开始摊销时间']) > cyStart else 0, axis=1) # and (stringToDate(row[u'开始摊销时间']) < CYend
df_rel[u'年初摊销结束标志']= df_rel.apply(lambda row: 1 if (convtPdTimeToDate(row[u'开始摊销时间']).replace(year=convtPdTimeToDate(row[u'开始摊销时间']).year+row[u'总摊销年限'])) < cyStart else 0, axis=1)
df_rel[u'年末摊销结束标志']= df_rel.apply(lambda row: 1 if (convtPdTimeToDate(row[u'开始摊销时间']).replace(year=convtPdTimeToDate(row[u'开始摊销时间']).year+row[u'总摊销年限'])) < cyEnd else 0, axis=1)
df_rel[u'每月摊销金额']    = df_rel.apply(lambda row: row[u'原始成本']/(row[u'总摊销年限']*12), axis=1)

df_rel[u'本年摊销月份']    = df_rel.apply(lambda row:  13 - month(convtPdTimeToStr(row[u'开始摊销时间'])) if row[u'本年新增标志'] == 1 else 
                                                     ( 0 if row[u'年初摊销结束标志'] == 1 else 
                                                     (  ( convtPdTimeToDate(row[u'开始摊销时间']).replace(year=convtPdTimeToDate(row[u'开始摊销时间']).year+row[u'总摊销年限']) ).month if row[u'年末摊销结束标志'] == 1 else 12 )) 
                                         ,axis=1)





df_rel[u'累计摊销月份']    = df_rel.apply(lambda row: min( monthTotal(cyEndStr) - monthTotal(convtPdTimeToStr(row[u'开始摊销时间']))+1 , row[u'总摊销年限']*12), axis=1)

df_rel[u'EY累计摊销额']    = df_rel.apply(lambda row: row[u'每月摊销金额']*row[u'累计摊销月份'], axis=1)
df_rel[u'DIFF']           = df_rel.apply(lambda row: row[u'EY累计摊销额']-row[u'累计摊销额'], axis=1)
df_rel[u'本年摊销金额']    = df_rel.apply(lambda row: row[u'每月摊销金额']*row[u'本年摊销月份'], axis=1)
df_rel[u'新增抽样标志']    = df_rel.apply(lambda row: row[u'每月摊销金额']*row[u'本年摊销月份'], axis=1)

df_rel.to_excel('res2.xlsx')





'''
list1 = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

list2 = [u'序号', u'资产编号', u'无形资产项目', u'无形资产分类', 
         u'总摊销年限', u'开始摊销时间', u'原始成本', u'累计摊销额']

list3 = [u'净值', u'本年新增标志', u'年初摊销结束标志', u'年末摊销结束标志', 
         u'每月摊销金额', u'本年摊销月份', u'累计摊销月份', 
         u'EY累计摊销额', u'DIFF', u'本年摊销金额', u'新增抽样标志']

path = 'C:\Workspace\AuditAutomation_L\L_workflow'

for root, dirs, files in os.walk(path):
    for file in files:
        if file.decode('gbk').find('~$') == -1 and file.decode('gbk').find('PBC') != -1: # eliminate temp excel files
            PBC = os.path.join(root,file.decode('gbk'))
        if file.decode('gbk').find('~$') == -1 and file.decode('gbk').find('WP') != -1:
            WP = os.path.join(root,file.decode('gbk'))

wb = load_workbook(WP)
ws = wb.get_sheet_by_name('L120')


净值            =原始成本-累计摊销额

本年新增标志     =if 开始摊销时间 in 本年年初~本年年末 then 1 else 0

年初摊销结束标志 =if 开始摊销时间+总摊销年限>本年年初 then 1 else 0
年末摊销结束标志 =if 开始摊销时间+总摊销年限>本年年末 then 1 else 0
每月摊销金额     =原始成本/总摊销年限*12
本年摊销月份     = if (本年新增标志 =1) then 13-month(开始摊销时间)
	               else if (本年新增标志 =0 &&年初摊销结束标志=1) then 0
	               else if (本年新增标志 =0 &&年末摊销结束标志=1) then 13 -month(开始摊销时间+总摊销年限)
	               else 12
累计摊销月份  = min(month(本年年末-开始摊销时间+1),总摊销年限*12)

EY累计摊销额  = 每月摊销金额*累计摊销月份

DIFF          = EY累计摊销额-累计摊销额
本年摊销金额  = 每月摊销金额*本年摊销月份
新增抽样标志  = if 本年新增标志=1&&阈值&抽样比例(pending)










column_index_from_string('T') #= '20'
for i in range(11,20):
    print ws[10][i].column,ws[10][i].value,' '*(11-len(ws[10][i].value)),' = ',ws[11][i].value
#for cell in ws[11]:
#    print cell.column, cell.value
    
  
    
    
    
    

L 本年新增标志         =  =IF(F11*1>BG!$B$5,1,0)
M 年初摊销结束标志      =  =IF((F11+E11*365)<($B$4-365),1,0)
N 本年处置前摊销结束标志=  =IF(M11=1,0,IF((F11+E11*365)<J11,1,0))
O 每月摊销金额        =  =IFERROR(G11/(E11*12),0)
P 本年摊销月份        =  =IF(M11=1,0,IF(L11=1,MONTH(J11)-MONTH(F11),IF(N11=1,MONTH(F11),MONTH(J11)-1)))
Q 累计摊销月份        =  =MIN((YEAR(J11)-YEAR(F11))*12+MONTH(J11)-MONTH(F11),E11*12)
R 本年摊销金额        =  =P11*O11
S EY累计摊销金额      =  =Q11*O11
T DIFF          =  =S11-H11
'''    
