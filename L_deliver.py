# -*- coding: utf-8 -*-
"""
Created on Wed Sep 27 10:04:59 2017

@author: Maddox.Meng
"""

'''L-line'''

import os
from openpyxl import load_workbook
import pandas as pd
#from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime  
import tkFileDialog

default_dir = r"%USERPROFILE%\Desktop"  # 设置默认打开目录
PBC = tkFileDialog.askopenfilename(title=u"选择文件",
                                     initialdir=(os.path.expanduser(default_dir)))

WP = tkFileDialog.askopenfilename(title=u"选择文件",
                                     initialdir=(os.path.expanduser(default_dir)))


'''========================================================================='''
def stringToDate(string): 
    #example '2013-07-22 09:44:15+00:00' 
    dt = datetime.datetime.strptime(string, "%Y-%m-%d") 
    return dt

def intToDays(year):
    return datetime.timedelta(days=365*int(year))

def month(date):
    return int(date.split('-')[1])

def year(date):
    return int(date.split('-')[0])

def monthTotal(date):
    return year(date)*12 + month(date)

'''处理当pandas读到日期格式为pandas._libs.tslib.Timestamp的cell，转成string（）'''
def convtPdTimeToStr(time): # datetime.date(2012, 9, 18)
    if isinstance(time, pd._libs.tslib.Timestamp) is True:
        dateStr = time.date().strftime('%Y-%m-%d')
    elif isinstance(time, unicode) is True or isinstance(time, str):
        dateStr = time
    else:
        dateStr = 'date type is %s' % type(time)
    return dateStr  

def convtPdTimeToDate(time):
    if isinstance(time, pd._libs.tslib.Timestamp) is True:
        dateStr = time.date()
    elif isinstance(time, unicode) is True or isinstance(time, str):
        dateStr = stringToDate(time)
    else:
        dateStr = 'date type is %s' % type(time)
    return dateStr 

#if datetime.date(2016,12,31) > stringToDate('2015-12-31'):
#    print 'yes'

path = 'C:\Workspace\AuditAutomation_L\L_workflow'
#CYstart = datetime.date(2016,12,31)#'2016-12-31'
#CYend   = datetime.date(2015,12,31)#'2015-12-31'

cyStart = stringToDate('2015-12-31')#'2016-12-31'
cyStartStr = '2015-12-31'
cyEnd   = stringToDate('2016-12-31')#'2015-12-31'
cyEndStr = '2016-12-31'
print cyStart + datetime.timedelta(days=365*5)

    
    
    
xl = pd.ExcelFile(PBC)
wb = load_workbook(WP)


from Tkinter import *
import Tkinter as tk
root = tk.Tk()
scrollbar = Scrollbar(root)
scrollbar.pack( side = RIGHT, fill=Y )
text = Text(root, yscrollcommand = scrollbar.set)
text.configure(font=("微软雅黑", 12))
scrollbar.config( command = text.yview )
def scrollwheel(event):
    text.yview_scroll(-1*(event.delta/120), "units")
text.bind_all('<MouseWheel>',scrollwheel)


#找到“序号” 

'''如果整个没找到，再打印error，怎么做到？'''
def getCoorByCellValue_pd(x_min, x_max, y_min, y_max, value):
     for x in range(x_min, x_max+1):
        for y in range(y_min, y_max+1):
            if df.iloc[x,y] == value:
                return x,y
            #else: print "ERROE: Couldn't find cell value %s" %value
           
def getCoorByCellValue_opx(x_min, x_max, y_min, y_max, value):
     for x in range(x_min, x_max+1):
        for y in range(y_min, y_max+1):
            if ws[x][y].value == value:
                return x,y                
            #else: print "ERROE: Couldn't find cell value %s" %value
            
def tkMessage(textWidget, textContent):
    textWidget.insert(INSERT, textContent)
    textWidget.update()
    textWidget.see('end')    
    
    
    
def addCalculatedCol(dataframe):
    #df_rel['sum'] = df_rel.apply(lambda x: x.sum(), axis=1)
    #df_rel['净值'] = df_rel.apply(lambda row: row[u'原始成本']-row[u'累计摊销额'], axis=1)
    
    dataframe[u'净值']            = dataframe.apply(lambda row: row[u'原始成本']-row[u'累计摊销额'], axis=1)
    dataframe[u'本年新增标志']    = dataframe.apply(lambda row: 1 if stringToDate(row[u'开始摊销时间']) > cyStart else 0, axis=1) # and (stringToDate(row[u'开始摊销时间']) < CYend
    dataframe[u'年初摊销结束标志']= dataframe.apply(lambda row: 1 if (stringToDate(row[u'开始摊销时间'])  +  intToDays(row[u'总摊销年限'])) < cyStart else 0, axis=1)
    dataframe[u'年末摊销结束标志']= dataframe.apply(lambda row: 1 if (stringToDate(row[u'开始摊销时间'])  +  intToDays(row[u'总摊销年限'])) < cyEnd else 0, axis=1)
    dataframe[u'每月摊销金额']    = dataframe.apply(lambda row: row[u'原始成本']/(row[u'总摊销年限']*12), axis=1)
    
    dataframe[u'本年摊销月份']    = dataframe.apply(lambda row:  13 - month(convtPdTimeToStr(row[u'开始摊销时间'])) if row[u'本年新增标志'] == 1 else 
                                                         ( 0 if row[u'年初摊销结束标志'] == 1 else 
                                                         (  ( convtPdTimeToDate(row[u'开始摊销时间']).replace(year=convtPdTimeToDate(row[u'开始摊销时间']).year+row[u'总摊销年限']) ).month if row[u'年末摊销结束标志'] == 1 else 12 )) 
                                             ,axis=1)
    
    dataframe[u'累计摊销月份']    = dataframe.apply(lambda row: min( monthTotal(cyEndStr) - monthTotal(row[u'开始摊销时间']), row[u'总摊销年限']*12), axis=1)
    
    dataframe[u'EY累计摊销额']    = dataframe.apply(lambda row: row[u'每月摊销金额']*row[u'累计摊销月份'], axis=1)
    dataframe[u'DIFF']            = dataframe.apply(lambda row: row[u'EY累计摊销额']-row[u'累计摊销额'], axis=1)
    dataframe[u'本年摊销金额']    = dataframe.apply(lambda row: row[u'每月摊销金额']*row[u'本年摊销月份'], axis=1)
    dataframe[u'新增抽样标志']    = dataframe.apply(lambda row: 'Y' if (row[u'本年新增标志'] == 1 and row[u'原始成本'] > 10000) else '', axis=1)
    
    #dataframe.to_excel('res.xlsx')    


logfile = open('logs.txt', 'w+')

for sheet_name in xl.sheet_names:#PBC的每个sheet L110-无形资产清单  L120-无形资产处置清单 
    print sheet_name
    df = pd.read_excel(open(PBC,'rb'), sheetname=sheet_name, header = None)
    
    logfile.write('  | -- '+sheet_name.encode('utf8')+'\n')
    tkMessage(text, '  | -- '+sheet_name+'\n')   
    
    
    #找到“序号” 
    x_max = df.shape[0]-1
    y_max = df.shape[1]- 1
    x_min = getCoorByCellValue_pd(0, x_max, 0, y_max, u'序号')[0]

                               
    #df_rel = df.iloc[x_min:df.shape[0], y_min:df.shape[1]] 
    #设置“序号”那行为header           
    df_rel = pd.read_excel(open(PBC,'rb'), sheetname=sheet_name, header = x_min)
    
    '''添加计算列'''
    addCalculatedCol(df_rel)
        
    '''写入WP'''
    '''我决定最佳方案是按列写入，所以先按列读取数据'''

         
     #在WP众多sheets里选出带LL110、L120的sheet
    for sheetname_WP in wb.sheetnames:
        
        if sheetname_WP in sheet_name:
            print sheetname_WP, "in", sheet_name
            ws = wb.get_sheet_by_name(sheetname_WP)  #打开它为ws

    rownumber = getCoorByCellValue_opx(1, ws.max_row, 0, ws.max_column-1, u'序号')[0]        

    logfile.write('    | -- WRITING FOLLOWING COLUMN\n')
    tkMessage(text, '    | -- WRITING FOLLOWING COLUMN\n')
                
    for header in list(df_rel):
        
        logfile.write('    | -- '+header.encode('utf8')+'\n')
        tkMessage(text, '    | -- '+header+'\n')
        text.pack()

            
        for cell in ws[rownumber]:
            if cell.value == header:
                
                print 'cell.vaue = header =',cell.value,cell.row,cell.column
                data_col = df_rel[header]
                data_wp_colNo = column_index_from_string(cell.column) - 1
                
                i=0
                y = data_wp_colNo
                #for x in range(rownumber+1, ws.max_row+1):
                for x in range(rownumber+1, rownumber+1+len(data_col)):    
                    
                    ws[x][y].value = data_col[i]
                    
                    i+=1
                    
wb.save(WP)           

tkMessage(text, 'Successfully!\n') 
root.mainloop()
                
            